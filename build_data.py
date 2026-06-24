# -*- coding: utf-8 -*-
"""Build data.js (dual-track: 物理类 + 历史类) for the app."""
import openpyxl, json
import os.path as _p

_ROOT = _p.dirname(_p.abspath(__file__))
_DIR = _p.join(_ROOT, 'source-data')

YEARS = [2023, 2024, 2025]

# 每个科类: key -> (显示名, 文件路径模板, 批次线"科类"标签)
TRACKS = {
    'wuli': ('物理类', _p.join(_DIR, '重庆{y}物理类.xlsx'), '物理类'),
    'lishi': ('历史类', _p.join(_DIR, '历史', '重庆{y}历史类.xlsx'), '历史类'),
}

# province centers from geojson meta
geo = json.load(open(_p.join(_ROOT, 'geo_meta.json'), encoding='utf-8'))
PROV = {k: {'n': v['name'], 'c': v['center']}
        for k, v in geo.items() if len(k) == 2 and v['center']}


def norm_tier(t):
    return str(t) if t else '普通院校'


def build_track(tmpl, pici_label):
    schools = {}
    lines = {}
    for yr in YEARS:
        path = tmpl.format(y=yr)
        if not _p.exists(path):
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        # batch lines
        ws = wb['批次线']
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == pici_label:
                lines[str(yr)] = {'tk': row[1], 'tkr': row[2], 'bk': row[3], 'bkr': row[4]}
                break
        # school ranking -> overall rank + admitted-count
        rankmap = {}
        ws = wb['院校排名']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            rankmap[str(row[1])] = {'r': row[0], 'nc': row[7]}
        # all-major detail
        ws = wb['全部专业明细']
        permajor, meta = {}, {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            code = str(row[0]) if row[0] is not None else None
            score, rank = row[4], row[5]
            if code is None or score is None or rank is None:
                continue
            permajor.setdefault(code, []).append([str(row[3]), int(score), int(rank)])
            meta[code] = (row[1], norm_tier(row[6]))
        for code, majors in permajor.items():
            majors.sort(key=lambda x: -x[1])
            name, tier = meta[code]
            scores = [m[1] for m in majors]; ranks = [m[2] for m in majors]
            ystat = {'mx': max(scores), 'mxr': min(ranks),
                     'mn': min(scores), 'mnr': max(ranks), 'nc': len(majors)}
            if code in rankmap:
                ystat['r'] = rankmap[code]['r']
            s = schools.setdefault(code, {'c': code, 'p': code[:2], 'y': {}, 'm': {}})
            s['n'] = name; s['t'] = tier
            s['y'][str(yr)] = ystat
            s['m'][str(yr)] = majors
        wb.close()
    school_list = [s for s in schools.values() if s['p'] in PROV]
    school_list.sort(key=lambda s: min((v['mnr'] for v in s['y'].values()), default=10**9))
    present_years = [y for y in YEARS if str(y) in lines or any(str(y) in s['y'] for s in school_list)]
    return {'years': present_years or YEARS, 'lines': lines, 'schools': school_list}


tracks = {}
for key, (name, tmpl, label) in TRACKS.items():
    t = build_track(tmpl, label)
    if t['schools']:
        t['name'] = name
        tracks[key] = t
        print(f'{name}: schools {len(t["schools"])}, years {t["years"]}')

out = {'prov': PROV, 'tracks': tracks}
js = 'window.GK=' + json.dumps(out, ensure_ascii=False, separators=(',', ':')) + ';'
open(_p.join(_ROOT, 'data.js'), 'w', encoding='utf-8').write(js)
print('data.js size(KB):', round(_p.getsize(_p.join(_ROOT, 'data.js')) / 1024, 1))
